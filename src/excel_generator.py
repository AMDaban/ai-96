from openpyxl import Workbook
from . import constants
from openpyxl.styles import Alignment


def create_result(state, result_file_name):
    result_workbook = Workbook()
    result_sheet = result_workbook.active

    max_number_of_filled = 0
    for day in state:
        for slot in state.get(day):
            max_number_of_filled = max(max_number_of_filled, len(state.get(day).get(slot)))

    result_sheet.append(
        tuple([" "]) + tuple(constants.day_time_slots)
    )

    week_days_pointer = 2
    for day in constants.week_days_complete:
        result_sheet.merge_cells(
            start_row=week_days_pointer,
            end_row=week_days_pointer + max_number_of_filled - 1,
            start_column=1,
            end_column=1
        )
        result_sheet.cell(week_days_pointer, 1).value = day
        week_days_pointer += max_number_of_filled

    vertical_offset = 2
    for day in state:
        col_number = 2
        for slot in state.get(day):
            index = 0
            for item in state.get(day).get(slot):
                result_sheet.cell(row=vertical_offset + index, column=col_number).value = to_string(item)
                index += 1
            col_number += 1
        vertical_offset += max_number_of_filled

    for col in ["A", "B", "C", "D", "E"]:
        result_sheet.column_dimensions[col].width = constants.result_file_column_width

    for i in range(1, 40):
        result_sheet.cell(i, 1).alignment = Alignment(horizontal='center', vertical='center')

    for j in range(1, 10):
        result_sheet.cell(1, j).alignment = Alignment(horizontal='center', vertical='center')

    for i in range(2, 40):
        for j in range(2, 10):
            result_sheet.cell(i, j).alignment = Alignment(horizontal='left', vertical='center')

    result_workbook.save(result_file_name)


def to_string(item):
    return str(str(item[0]) + " " + item[1] + " " + item[2])
