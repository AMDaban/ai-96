from openpyxl import load_workbook
from functools import reduce

professor_skills_map = None
lessons_list = None
classes_set = None
subject_set = None
free_times_map = None


def load_resources(professor_skills, classes, subjects, free_times):
    load_professor_skills(professor_skills_file_name=professor_skills)
    load_classes(classes_file_name=classes)
    load_subjects(subject_file_name=subjects)
    load_free_times(free_times_file_name=free_times)


def load_professor_skills(professor_skills_file_name):
    global professor_skills_map, lessons_list

    prof_skills = load_workbook(professor_skills_file_name)
    prof_skills_sheet = prof_skills.active

    lessons = prof_skills_sheet.iter_rows(
        min_row=1,
        min_col=2,
        max_row=1,
        max_col=prof_skills_sheet.max_column
    )

    lessons_list = list(
        map(
            lambda x: x.value,
            reduce(
                lambda x, y: x.append(y),
                [row for row in lessons]
            )
        )
    )

    professors = prof_skills_sheet.iter_rows(
        min_row=2,
        min_col=1,
        max_row=prof_skills_sheet.max_row,
        max_col=prof_skills_sheet.max_column
    )

    professor_skills_map = {
        professor[0].value: get_skill_map(professor[1:]) for professor in professors
    }


def get_skill_map(professor_skills):
    skills_map = {}

    for i in range(0, len(professor_skills)):
        skills_map.update({
            lessons_list[i]: professor_skills[i].value
        })

    return skills_map


def get_professor_skills(professor):
    global professor_skills_map

    if professor_skills_map is not None:
        return professor_skills_map.get(professor, None)
    else:
        return None


def get_masters_for(subject):
    if professor_skills_map is None:
        return set({})

    temp_subject_masters = set({})
    for professor in professor_skills_map:
        if professor_skills_map.get(professor).get(subject) == 1:
            temp_subject_masters.add(professor)

    return temp_subject_masters


def get_professors_free_time_count():
    professors = set([professor for professor in professor_skills_map])

    result = {}

    for professor in professors:
        professor_week = free_times_map.get(professor)
        temp_professor_free_time_count = 0
        for day in professor_week:
            for slot in professor_week.get(day):
                temp_professor_free_time_count += professor_week.get(day).get(slot)

        result.update({
            professor: temp_professor_free_time_count
        })

    return result


def load_classes(classes_file_name):
    global classes_set

    classes_work_book = load_workbook(classes_file_name)
    classes_sheet = classes_work_book.active

    classes = classes_sheet.iter_rows(
        min_row=1,
        min_col=1,
        max_row=classes_sheet.max_row,
        max_col=classes_sheet.max_column
    )

    classes_set = set(
        list(
            map(
                lambda x: x.value,
                reduce(
                    (lambda x, y: x.append(y)),
                    [row for row in classes]
                )
            )
        )
    )


def get_classes():
    return classes_set


def load_subjects(subject_file_name):
    global subject_set

    subjects_work_book = load_workbook(subject_file_name)
    subject_sheet = subjects_work_book.active

    subjects = subject_sheet.iter_rows(
        min_row=1,
        min_col=1,
        max_row=subject_sheet.max_row,
        max_col=subject_sheet.max_column
    )

    subject_set = set(
        list(
            map(
                lambda x: x.value,
                reduce(
                    lambda x, y: x.append(y),
                    [row for row in subjects]
                )
            )
        )
    )


def get_subjects():
    return subject_set


def load_free_times(free_times_file_name):
    global free_times_map

    free_times_work_book = load_workbook(free_times_file_name)

    free_times_map = {}

    for professor_free_time_sheet in free_times_work_book:
        sheet_rows = professor_free_time_sheet.iter_rows(
            min_row=2,
            min_col=1,
            max_row=professor_free_time_sheet.max_row,
            max_col=professor_free_time_sheet.max_column
        )

        professor_week_free_time = {
            row[0].value: get_professor_day_free_time(row[1:]) for row in sheet_rows
        }

        free_times_map.update({
            professor_free_time_sheet.title: professor_week_free_time
        })


day_time_slots = ["8-10", "10-12", "14-16", "16-18"]


def get_professor_day_free_time(day):
    professor_day_free_time = {}

    for i in range(0, len(day)):
        professor_day_free_time.update({
            day_time_slots[i]: day[i].value
        })

    return professor_day_free_time


def get_professor_free_time(professor):
    global free_times_map

    if free_times_map is not None:
        return free_times_map.get(professor, None)
    else:
        return None

